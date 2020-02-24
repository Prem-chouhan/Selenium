import openpyxl

path = "/home/admin-1/prem/write.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for r in range(1, 5):
    for c in range(1, 17):
        sheet.cell(row=r, column=c).value = "welcome"

workbook.save(path)
